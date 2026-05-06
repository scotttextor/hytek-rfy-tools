"""Build a multi-job stick-comparison spreadsheet from staged data.

Reads tmp_detailer_test/multi-job/<jobnum>/{*.detailer-ref.xml, *.codec.xml, meta.json}
and assembles ONE workbook with:
  - Sheet "_Summary": per-job ops/match counts + gap aggregation
  - Sheet "<jobnum>": per-frame per-stick side-by-side ops with diff highlighting
"""
import os
import json
import xml.etree.ElementTree as ET
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r"C:\Users\Scott\CLAUDE CODE\hytek-rfy-tools\tmp_detailer_test\multi-job"
OUT = r"C:\Users\Scott\CLAUDE CODE\hytek-rfy-tools\tmp_detailer_test\MULTI-JOB-COMPARE.xlsx"


def parse_inner_xml(path):
    tree = ET.parse(path)
    root = tree.getroot()
    out = {}
    for plan in root.iter("plan"):
        for frame in plan.findall("frame"):
            fname = frame.get("name", "?")
            stick_map = {}
            for stick in frame.findall("stick"):
                sname = stick.get("name", "?")
                ops = []
                tooling = stick.find("tooling")
                if tooling is not None:
                    for op in tooling:
                        ops.append({
                            "tag": op.tag, "type": op.get("type", "?"),
                            "pos": op.get("pos"),
                            "start": op.get("startPos"), "end": op.get("endPos"),
                        })
                stick_map[sname] = ops
            out[fname] = stick_map
    return out


def op_to_str(op):
    tag = op["tag"].replace("-tool", "")
    t = op["type"]
    if op.get("pos") is not None:
        return f"{t} @ {op['pos']} ({tag})"
    if op.get("start") is not None and op.get("end") is not None:
        return f"{t} {op['start']}..{op['end']} ({tag})"
    return f"{t} ({tag})"


def op_key(op):
    p = op.get("pos") or op.get("start") or "0"
    try: p = float(p)
    except Exception: p = 0
    return (op["type"], op["tag"], p)


def pair_ops(ref_ops, codec_ops):
    """Pair Detailer ops with codec ops by (type, tag, pos within 5mm)."""
    codec_pool = list(codec_ops)
    paired = []
    matched = 0
    for r_op in ref_ops:
        best_idx = -1
        for i, c_op in enumerate(codec_pool):
            if c_op["type"] == r_op["type"] and c_op["tag"] == r_op["tag"]:
                rp = float(r_op.get("pos") or r_op.get("start") or 0)
                cp = float(c_op.get("pos") or c_op.get("start") or 0)
                if abs(rp - cp) < 5:
                    best_idx = i; break
        if best_idx >= 0:
            paired.append((r_op, codec_pool.pop(best_idx), "match"))
            matched += 1
        else:
            paired.append((r_op, None, "missing"))
    for c_op in codec_pool:
        paired.append((None, c_op, "extra"))
    return paired, matched


def main():
    if not os.path.isdir(BASE):
        print(f"ERROR: {BASE} not found")
        return

    # Discover jobs
    jobs = []
    for jobnum in sorted(os.listdir(BASE)):
        jp = os.path.join(BASE, jobnum)
        if not os.path.isdir(jp): continue
        meta_path = os.path.join(jp, "meta.json")
        if not os.path.isfile(meta_path): continue
        with open(meta_path) as f:
            meta = json.load(f)
        ref_xml = next((os.path.join(jp, f) for f in os.listdir(jp) if f.endswith(".detailer-ref.xml")), None)
        codec_xml = next((os.path.join(jp, f) for f in os.listdir(jp) if f.endswith(".codec.xml")), None)
        if ref_xml and codec_xml:
            jobs.append({"jobnum": jobnum, "meta": meta, "ref_xml": ref_xml, "codec_xml": codec_xml})

    print(f"Building spreadsheet for {len(jobs)} jobs:")
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    H_FILL = PatternFill("solid", start_color="305496")
    SECT_FONT = Font(name="Arial", bold=True, size=11)
    SECT_FILL = PatternFill("solid", start_color="D9E1F2")
    BODY = Font(name="Arial", size=10)
    DIFF_FONT = Font(name="Arial", color="C00000", bold=True)
    DIFF_FILL = PatternFill("solid", start_color="FFCCCC")
    MISS_FILL = PatternFill("solid", start_color="FFE699")
    EXTRA_FILL = PatternFill("solid", start_color="FFC7CE")
    POS_FILL = PatternFill("solid", start_color="FFF2CC")
    MATCH_FONT = Font(name="Arial", color="385723")
    POS_FONT = Font(name="Arial", italic=True, color="806000")
    LINK_FONT = Font(name="Arial", color="0563C1", underline="single", size=10)
    thin = Side(border_style="thin", color="BFBFBF")
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ==== Summary sheet (built last after we know totals) ====
    summary_data = []
    aggregate_op_gaps = Counter()  # by (op_type, status) → count

    for job in jobs:
        jobnum = job["jobnum"]
        meta = job["meta"]
        plan_name = meta["planName"]

        print(f"  {jobnum} {plan_name}...")
        ref = parse_inner_xml(job["ref_xml"])
        codec = parse_inner_xml(job["codec_xml"])

        ws = wb.create_sheet(title=jobnum)

        # Header info block
        ws.cell(row=1, column=1, value=f"{jobnum} — {meta['builder']} — {plan_name}").font = Font(name="Arial", bold=True, size=14)
        ws.merge_cells("A1:F1")
        for i, (label, val) in enumerate([
            ("Input XML (Y:):", meta["xmlPath"]),
            ("Detailer ref RFY (Y:):", meta["refRfyPath"]),
            ("Job dir (Y:):", meta.get("jobDir", "")),
        ], start=3):
            ws.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
            c = ws.cell(row=i, column=2, value=val)
            c.font = LINK_FONT
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
            if val.startswith(("Y:", "C:")):
                c.hyperlink = "file:///" + val.replace("\\", "/")

        # Table header
        table_start = 7
        for col, h in enumerate(["Frame", "Stick / D:codec", "Op #", "Detailer (ref)", "Codec (ours)", "Match?"], 1):
            c = ws.cell(row=table_start, column=col, value=h)
            c.font = H_FONT; c.fill = H_FILL; c.border = BOX
            c.alignment = Alignment(horizontal="center")
        row = table_start + 1

        total_d = 0; total_c = 0; total_m = 0
        op_gap_counts = Counter()  # per-op-type missing/extra for THIS job

        for fname in ref.keys():
            ref_sticks = ref[fname]
            codec_sticks = codec.get(fname, {})
            all_sticks = list(ref_sticks.keys())
            for s in codec_sticks.keys():
                if s not in all_sticks: all_sticks.append(s)

            # Frame header
            c = ws.cell(row=row, column=1, value=f"FRAME {fname}")
            c.font = SECT_FONT; c.fill = SECT_FILL; c.border = BOX
            for col in range(2, 7):
                ws.cell(row=row, column=col).fill = SECT_FILL
                ws.cell(row=row, column=col).border = BOX
            row += 1

            for sname in all_sticks:
                r_ops = sorted(ref_sticks.get(sname, []), key=op_key)
                c_ops = sorted(codec_sticks.get(sname, []), key=op_key)
                total_d += len(r_ops); total_c += len(c_ops)

                # Stick subhead
                stick_label = f"{sname}  (D:{len(r_ops)} / C:{len(c_ops)})"
                c2 = ws.cell(row=row, column=2, value=stick_label)
                c2.font = Font(name="Arial", bold=True, size=10); c2.border = BOX
                c2.fill = PatternFill("solid", start_color="F2F2F2")
                for col in [1, 3, 4, 5, 6]:
                    ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="F2F2F2")
                    ws.cell(row=row, column=col).border = BOX
                row += 1

                paired, matched = pair_ops(r_ops, c_ops)
                total_m += matched

                for op_idx, (r_op, c_op, status) in enumerate(paired, 1):
                    ws.cell(row=row, column=1).border = BOX
                    ws.cell(row=row, column=2).border = BOX
                    ws.cell(row=row, column=3, value=op_idx).border = BOX
                    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                    ref_str = op_to_str(r_op) if r_op else ""
                    codec_str = op_to_str(c_op) if c_op else ""
                    rc = ws.cell(row=row, column=4, value=ref_str); rc.font = BODY; rc.border = BOX
                    cc = ws.cell(row=row, column=5, value=codec_str); cc.font = BODY; cc.border = BOX

                    if status == "match":
                        rp = (r_op.get("pos") or f"{r_op.get('start')}..{r_op.get('end')}")
                        cp = (c_op.get("pos") or f"{c_op.get('start')}..{c_op.get('end')}")
                        if rp != cp:
                            ws.cell(row=row, column=6, value="~match (pos diff)").font = POS_FONT
                            rc.fill = POS_FILL; cc.fill = POS_FILL
                        else:
                            ws.cell(row=row, column=6, value="match").font = MATCH_FONT
                    elif status == "missing":
                        rc.fill = MISS_FILL; cc.fill = MISS_FILL
                        cc.value = "(missing — codec doesn't emit)"; cc.font = DIFF_FONT
                        ws.cell(row=row, column=6, value="MISSING").font = DIFF_FONT
                        ws.cell(row=row, column=6).fill = DIFF_FILL
                        op_gap_counts[(r_op["type"], "missing")] += 1
                        aggregate_op_gaps[(r_op["type"], "missing")] += 1
                    else:
                        rc.fill = EXTRA_FILL; cc.fill = EXTRA_FILL
                        rc.value = "(extra — codec wrongly emits)"; rc.font = DIFF_FONT
                        ws.cell(row=row, column=6, value="EXTRA").font = DIFF_FONT
                        ws.cell(row=row, column=6).fill = DIFF_FILL
                        op_gap_counts[(c_op["type"], "extra")] += 1
                        aggregate_op_gaps[(c_op["type"], "extra")] += 1

                    ws.cell(row=row, column=6).border = BOX
                    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
                    row += 1

                row += 1  # blank

        # Job totals at bottom
        row += 1
        ws.cell(row=row, column=1, value="JOB TOTALS").font = Font(name="Arial", bold=True, size=12)
        row += 1
        for label, value in [
            ("Detailer ops:", total_d),
            ("Codec ops:", total_c),
            ("Matched:", total_m),
            ("Match %:", f"{100*total_m/total_d:.1f}%" if total_d else "-"),
            ("Codec MISSING:", total_d - total_m),
            ("Codec EXTRA:", total_c - total_m),
        ]:
            ws.cell(row=row, column=4, value=label).font = Font(name="Arial", bold=True)
            ws.cell(row=row, column=5, value=value)
            row += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 6
        ws.column_dimensions["D"].width = 36
        ws.column_dimensions["E"].width = 36
        ws.column_dimensions["F"].width = 20
        ws.freeze_panes = f"A{table_start + 1}"

        summary_data.append({
            "jobnum": jobnum, "builder": meta["builder"], "plan": plan_name,
            "detailer_ops": total_d, "codec_ops": total_c, "matched": total_m,
            "missing": total_d - total_m, "extra": total_c - total_m,
            "match_pct": (100 * total_m / total_d) if total_d else 0,
            "op_gaps": op_gap_counts,
        })

    # ==== Summary sheet ====
    s = wb.create_sheet(title="_Summary", index=0)
    s.cell(row=1, column=1, value="HYTEK Codec vs Detailer — multi-job comparison").font = Font(name="Arial", bold=True, size=14)
    s.merge_cells("A1:H1")

    s.cell(row=3, column=1, value="Per-job match summary").font = Font(name="Arial", bold=True, size=12)
    headers = ["Job", "Builder", "Plan", "Detailer ops", "Codec ops", "Matched", "Missing", "Extra", "Match %"]
    for col, h in enumerate(headers, 1):
        c = s.cell(row=4, column=col, value=h)
        c.font = H_FONT; c.fill = H_FILL; c.border = BOX
        c.alignment = Alignment(horizontal="center")

    row = 5
    grand_d = grand_c = grand_m = 0
    for sd in summary_data:
        s.cell(row=row, column=1, value=sd["jobnum"]).border = BOX
        s.cell(row=row, column=2, value=sd["builder"]).border = BOX
        s.cell(row=row, column=3, value=sd["plan"]).border = BOX
        s.cell(row=row, column=4, value=sd["detailer_ops"]).border = BOX
        s.cell(row=row, column=5, value=sd["codec_ops"]).border = BOX
        s.cell(row=row, column=6, value=sd["matched"]).border = BOX
        s.cell(row=row, column=7, value=sd["missing"]).border = BOX
        s.cell(row=row, column=8, value=sd["extra"]).border = BOX
        s.cell(row=row, column=9, value=f"{sd['match_pct']:.1f}%").border = BOX
        if sd["match_pct"] < 75:
            s.cell(row=row, column=9).fill = DIFF_FILL
            s.cell(row=row, column=9).font = DIFF_FONT
        elif sd["match_pct"] < 85:
            s.cell(row=row, column=9).fill = POS_FILL
        else:
            s.cell(row=row, column=9).font = MATCH_FONT
        grand_d += sd["detailer_ops"]; grand_c += sd["codec_ops"]; grand_m += sd["matched"]
        row += 1

    # Grand total
    s.cell(row=row, column=1, value="TOTAL").font = SECT_FONT
    s.cell(row=row, column=4, value=grand_d).font = SECT_FONT
    s.cell(row=row, column=5, value=grand_c).font = SECT_FONT
    s.cell(row=row, column=6, value=grand_m).font = SECT_FONT
    s.cell(row=row, column=7, value=grand_d - grand_m).font = SECT_FONT
    s.cell(row=row, column=8, value=grand_c - grand_m).font = SECT_FONT
    s.cell(row=row, column=9, value=f"{100*grand_m/grand_d:.1f}%" if grand_d else "-").font = SECT_FONT
    for col in range(1, 10):
        s.cell(row=row, column=col).fill = SECT_FILL
        s.cell(row=row, column=col).border = BOX
    row += 3

    # Aggregate op-gap pattern
    s.cell(row=row, column=1, value="Op-type gap pattern (across all jobs)").font = Font(name="Arial", bold=True, size=12)
    row += 1
    for col, h in enumerate(["Op type", "MISSING from codec", "EXTRA in codec"], 1):
        c = s.cell(row=row, column=col, value=h)
        c.font = H_FONT; c.fill = H_FILL; c.border = BOX
    row += 1
    op_types = sorted(set(t for (t, _), _ in aggregate_op_gaps.items()),
                       key=lambda t: -(aggregate_op_gaps.get((t, "missing"), 0) + aggregate_op_gaps.get((t, "extra"), 0)))
    for op_t in op_types:
        miss = aggregate_op_gaps.get((op_t, "missing"), 0)
        extr = aggregate_op_gaps.get((op_t, "extra"), 0)
        s.cell(row=row, column=1, value=op_t).border = BOX
        m_cell = s.cell(row=row, column=2, value=miss); m_cell.border = BOX
        e_cell = s.cell(row=row, column=3, value=extr); e_cell.border = BOX
        if miss > 0:
            m_cell.fill = MISS_FILL; m_cell.font = DIFF_FONT
        if extr > 0:
            e_cell.fill = EXTRA_FILL; e_cell.font = DIFF_FONT
        row += 1

    s.column_dimensions["A"].width = 14
    s.column_dimensions["B"].width = 26
    s.column_dimensions["C"].width = 22
    s.column_dimensions["D"].width = 14
    s.column_dimensions["E"].width = 14
    s.column_dimensions["F"].width = 12
    s.column_dimensions["G"].width = 12
    s.column_dimensions["H"].width = 10
    s.column_dimensions["I"].width = 12
    s.freeze_panes = "A5"

    wb.save(OUT)
    print(f"\nSaved: {OUT}")
    print(f"  Total Detailer ops across {len(summary_data)} jobs: {grand_d:,}")
    print(f"  Total Codec ops:    {grand_c:,}")
    print(f"  Matched:            {grand_m:,}")
    if grand_d:
        print(f"  Overall match %:    {100*grand_m/grand_d:.1f}%")


if __name__ == "__main__":
    main()
