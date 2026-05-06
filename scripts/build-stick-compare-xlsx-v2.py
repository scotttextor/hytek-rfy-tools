"""Side-by-side stick-comparison spreadsheet (v2).

Adds:
- Top header block showing Y: drive source-file paths
- Per-frame inline hyperlinks to the input XML on Y: drive
- Per-frame extracted XML snippet saved alongside (one file per frame)
"""
import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JOB_NAME = "HG260017 LOT 925 (42) STUNNING CRESCENT NARANGBA"
PLAN_NAME = "GF-LBW-70.075"
JOB_DIR = r"Y:\(17) 2026 HYTEK PROJECTS\CORAL HOMES\HG260017 LOT 925 (42) STUNNING CRESCENT NARANGBA"
INPUT_XML_Y = os.path.join(
    JOB_DIR, r"03 DETAILING\03 FRAMECAD DETAILER\01 XML OUTPUT",
    f"{JOB_NAME}-{PLAN_NAME}.xml"
)
REF_RFY_Y_DIR = os.path.join(JOB_DIR, r"06 MANUFACTURING\04 ROLLFORMER FILES\HG260017_SPLIT_2026-03-05")
REF_RFY_Y_PATH = os.path.join(REF_RFY_Y_DIR, f"HG260017_PK4-{PLAN_NAME}.rfy")

# Local artifacts (decrypted inner XMLs from earlier comparison)
LOCAL_BASE = r"C:\Users\Scott\CLAUDE CODE\hytek-rfy-tools\tmp_detailer_test"
DETAILER_INNER_XML = os.path.join(LOCAL_BASE, "COMPARE-2-detailer-now.xml")
CODEC_INNER_XML = os.path.join(LOCAL_BASE, "COMPARE-3-codec.xml")
DETAILER_RFY = os.path.join(LOCAL_BASE, "test_output.rfy")
PER_FRAME_DIR = os.path.join(LOCAL_BASE, "per-frame-xml")
OUT_PATH = os.path.join(LOCAL_BASE, "STICK-COMPARE-HG260017-GF-LBW-70.075.xlsx")


def parse_inner_xml(path):
    tree = ET.parse(path)
    root = tree.getroot()
    out = {}
    for plan in root.iter("plan"):
        for frame in plan.findall("frame"):
            fname = frame.get("name", "?")
            stick_map = {}
            for stick in frame.findall("stick"):
                sname = stick.get("name", "?")
                ops = []
                tooling = stick.find("tooling")
                if tooling is not None:
                    for op in tooling:
                        ops.append({
                            "tag": op.tag,
                            "type": op.get("type", "?"),
                            "pos": op.get("pos"),
                            "start": op.get("startPos"),
                            "end": op.get("endPos"),
                        })
                stick_map[sname] = ops
            out[fname] = stick_map
    return out


def extract_frame_xmls(inner_xml_path: str, out_dir: str, label: str) -> dict:
    """Extract each <frame> as its own .xml file. Returns {frame_name: file_path}."""
    os.makedirs(out_dir, exist_ok=True)
    tree = ET.parse(inner_xml_path)
    root = tree.getroot()
    paths = {}
    for plan in root.iter("plan"):
        for frame in plan.findall("frame"):
            fname = frame.get("name", "?")
            safe = "".join(c for c in fname if c.isalnum() or c in "-_") or "frame"
            out_path = os.path.join(out_dir, f"{safe}-{label}.xml")
            ET.ElementTree(frame).write(out_path, encoding="utf-8", xml_declaration=True)
            paths[fname] = out_path
    return paths


def op_to_str(op):
    tag = op["tag"].replace("-tool", "")
    t = op["type"]
    if op.get("pos") is not None:
        return f"{t} @ {op['pos']} ({tag})"
    if op.get("start") is not None and op.get("end") is not None:
        return f"{t} {op['start']}..{op['end']} ({tag})"
    return f"{t} ({tag})"


def op_key(op):
    p = op.get("pos") or op.get("start") or "0"
    try: p = float(p)
    except Exception: p = 0
    return (op["type"], op["tag"], p)


def main():
    print("Extracting per-frame XMLs (Detailer)...")
    detailer_frames = extract_frame_xmls(
        DETAILER_INNER_XML, os.path.join(PER_FRAME_DIR, "detailer"), "detailer")
    print(f"  saved {len(detailer_frames)} files to {PER_FRAME_DIR}/detailer/")

    print("Extracting per-frame XMLs (codec)...")
    codec_frames = extract_frame_xmls(
        CODEC_INNER_XML, os.path.join(PER_FRAME_DIR, "codec"), "codec")
    print(f"  saved {len(codec_frames)} files to {PER_FRAME_DIR}/codec/")

    ref = parse_inner_xml(DETAILER_INNER_XML)
    codec = parse_inner_xml(CODEC_INNER_XML)

    wb = Workbook()
    ws = wb.active
    ws.title = "HG260017 GF-LBW-70.075"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    section_font = Font(name="Arial", bold=True, size=11)
    section_fill = PatternFill("solid", start_color="D9E1F2")
    info_font = Font(name="Arial", bold=True, size=10)
    link_font = Font(name="Arial", color="0563C1", underline="single", size=10)
    diff_fill = PatternFill("solid", start_color="FFCCCC")
    diff_font = Font(name="Arial", color="C00000", bold=True)
    detailer_only_fill = PatternFill("solid", start_color="FFE699")
    codec_only_fill = PatternFill("solid", start_color="FFC7CE")
    body_font = Font(name="Arial", size=10)
    thin = Side(border_style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Top info block ----
    ws.cell(row=1, column=1, value="HYTEK Detailer vs Codec — operation-level comparison").font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells("A1:F1")

    info_rows = [
        ("Job:", JOB_NAME),
        ("Plan:", PLAN_NAME),
        ("Input XML (Y: drive):", INPUT_XML_Y),
        ("Detailer reference RFY (Y: drive):", REF_RFY_Y_PATH),
        ("Detailer fresh RFY (this PC):", DETAILER_RFY),
        ("Per-frame XMLs (this PC):", PER_FRAME_DIR),
    ]
    for i, (label, value) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=1, value=label).font = info_font
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        c = ws.cell(row=i, column=3, value=value)
        c.font = link_font if value.startswith(("Y:", "C:")) else body_font
        if value.startswith(("Y:", "C:")):
            # Build hyperlink — file:/// URI
            uri = "file:///" + value.replace("\\", "/")
            c.hyperlink = uri
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)

    # Legend
    legend_row = 3 + len(info_rows) + 1
    ws.cell(row=legend_row, column=1, value="Legend:").font = info_font
    ws.cell(row=legend_row, column=2, value="match").fill = PatternFill("solid", start_color="E2EFDA")
    ws.cell(row=legend_row, column=2).font = Font(name="Arial", color="385723")
    ws.cell(row=legend_row, column=3, value="~match (pos drift)").fill = PatternFill("solid", start_color="FFF2CC")
    ws.cell(row=legend_row, column=3).font = Font(name="Arial", italic=True, color="806000")
    ws.cell(row=legend_row, column=4, value="MISSING from codec").fill = detailer_only_fill
    ws.cell(row=legend_row, column=4).font = diff_font
    ws.cell(row=legend_row, column=5, value="EXTRA in codec").fill = codec_only_fill
    ws.cell(row=legend_row, column=5).font = diff_font

    # ---- Table headers ----
    table_start = legend_row + 2
    headers = ["Frame", "Stick / Frame XMLs", "Op #", "Detailer (ref)", "Codec (ours)", "Match?"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=table_start, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = box

    row = table_start + 1
    total_match = 0
    total_detailer = 0
    total_codec = 0

    for fname in ref.keys():
        ref_sticks = ref[fname]
        codec_sticks = codec.get(fname, {})
        all_stick_names = list(ref_sticks.keys())
        for s in codec_sticks.keys():
            if s not in all_stick_names:
                all_stick_names.append(s)

        # Frame header row with hyperlinks to per-frame XMLs
        frame_label = f"FRAME {fname}"
        c = ws.cell(row=row, column=1, value=frame_label)
        c.font = section_font; c.fill = section_fill; c.border = box

        # Stick column shows links to per-frame extracted XMLs
        d_xml = detailer_frames.get(fname)
        c_xml = codec_frames.get(fname)
        link_text = []
        if d_xml: link_text.append(f"Detailer XML: {os.path.basename(d_xml)}")
        if c_xml: link_text.append(f"Codec XML: {os.path.basename(c_xml)}")
        c2 = ws.cell(row=row, column=2, value=" | ".join(link_text))
        c2.font = link_font; c2.fill = section_fill; c2.border = box
        if d_xml:
            c2.hyperlink = "file:///" + d_xml.replace("\\", "/")

        # Reference URI in column 6 — points to the input XML on Y: drive
        c6 = ws.cell(row=row, column=6, value="Open input XML (Y:)")
        c6.font = link_font; c6.fill = section_fill; c6.border = box
        c6.hyperlink = "file:///" + INPUT_XML_Y.replace("\\", "/")

        for col in range(3, 6):
            ws.cell(row=row, column=col).fill = section_fill
            ws.cell(row=row, column=col).border = box

        row += 1

        for sname in all_stick_names:
            ref_ops = sorted(ref_sticks.get(sname, []), key=op_key)
            codec_ops = sorted(codec_sticks.get(sname, []), key=op_key)
            total_detailer += len(ref_ops)
            total_codec += len(codec_ops)

            stick_summary = f"{sname}  (D:{len(ref_ops)} / C:{len(codec_ops)})"
            c = ws.cell(row=row, column=1, value="").border = box
            c2 = ws.cell(row=row, column=2, value=stick_summary)
            c2.font = Font(name="Arial", bold=True, size=10); c2.border = box
            c2.fill = PatternFill("solid", start_color="F2F2F2")
            for col in range(3, 7):
                ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="F2F2F2")
                ws.cell(row=row, column=col).border = box
            row += 1

            codec_pool = list(codec_ops)
            paired_rows = []
            for r_op in ref_ops:
                best_idx = -1
                for i, c_op in enumerate(codec_pool):
                    if c_op["type"] == r_op["type"] and c_op["tag"] == r_op["tag"]:
                        rp = float(r_op.get("pos") or r_op.get("start") or 0)
                        cp = float(c_op.get("pos") or c_op.get("start") or 0)
                        if abs(rp - cp) < 5:
                            best_idx = i; break
                if best_idx >= 0:
                    matched = codec_pool.pop(best_idx)
                    paired_rows.append((r_op, matched, "match"))
                    total_match += 1
                else:
                    paired_rows.append((r_op, None, "missing"))
            for c_op in codec_pool:
                paired_rows.append((None, c_op, "extra"))

            for op_idx, (r_op, c_op, status) in enumerate(paired_rows, 1):
                ws.cell(row=row, column=1, value="").border = box
                ws.cell(row=row, column=2, value="").border = box
                ws.cell(row=row, column=3, value=op_idx).border = box
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

                ref_str = op_to_str(r_op) if r_op else ""
                codec_str = op_to_str(c_op) if c_op else ""
                rc = ws.cell(row=row, column=4, value=ref_str); rc.font = body_font; rc.border = box
                cc = ws.cell(row=row, column=5, value=codec_str); cc.font = body_font; cc.border = box

                if status == "match":
                    rp = (r_op.get("pos") or f"{r_op.get('start')}..{r_op.get('end')}")
                    cp = (c_op.get("pos") or f"{c_op.get('start')}..{c_op.get('end')}")
                    if rp != cp:
                        ws.cell(row=row, column=6, value="~match (pos diff)").font = Font(name="Arial", italic=True, color="806000")
                        rc.fill = PatternFill("solid", start_color="FFF2CC")
                        cc.fill = PatternFill("solid", start_color="FFF2CC")
                    else:
                        ws.cell(row=row, column=6, value="match").font = Font(name="Arial", color="385723")
                elif status == "missing":
                    cc.fill = detailer_only_fill
                    cc.value = "(missing — codec doesn't emit)"
                    cc.font = diff_font
                    rc.fill = detailer_only_fill
                    ws.cell(row=row, column=6, value="MISSING").font = diff_font
                    ws.cell(row=row, column=6).fill = diff_fill
                else:
                    rc.fill = codec_only_fill
                    rc.value = "(extra — codec wrongly emits)"
                    rc.font = diff_font
                    cc.fill = codec_only_fill
                    ws.cell(row=row, column=6, value="EXTRA").font = diff_font
                    ws.cell(row=row, column=6).fill = diff_fill

                ws.cell(row=row, column=6).border = box
                ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
                row += 1

            row += 1  # blank between sticks

    # Summary
    summary_row = row + 1
    ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(name="Arial", bold=True, size=12)
    summary_row += 1
    pairs = [
        ("Detailer ops:", total_detailer),
        ("Codec ops:", total_codec),
        ("Matched:", total_match),
        ("Match %:", f"{100*total_match/total_detailer:.1f}%" if total_detailer else "-"),
        ("Codec MISSING:", total_detailer - total_match),
        ("Codec EXTRA:", total_codec - total_match),
    ]
    for label, value in pairs:
        ws.cell(row=summary_row, column=4, value=label).font = Font(name="Arial", bold=True)
        ws.cell(row=summary_row, column=5, value=value)
        summary_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 22

    ws.freeze_panes = f"A{table_start + 1}"

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
    print(f"  Detailer ops: {total_detailer}")
    print(f"  Codec ops:    {total_codec}")
    print(f"  Match %:      {100*total_match/total_detailer:.1f}%")


if __name__ == "__main__":
    main()
