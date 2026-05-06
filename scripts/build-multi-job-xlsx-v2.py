"""Multi-job stick-comparison v2 with NEAREST-NEIGHBOUR matching.

v1 used first-fit within 5mm. That over-counts gaps: if codec emits an op of
the right type at 8mm off Detailer's, both sides count it (missing + extra).

v2: per (op_type, tag) group within a stick, optimally pair by nearest
position. Categorise outcomes:
  - exact          (≤ 0.5 mm) — green
  - drift-small    (≤ 5 mm)   — pale yellow (was "match" before)
  - drift-medium   (≤ 30 mm)  — orange (NEW — caught these as missing+extra in v1)
  - drift-large    (> 30 mm)  — red but TYPE matches
  - missing        — no codec op of this type at all
  - extra          — no Detailer op of this type at all

This shows whether the gaps are TRULY missing rules or just position drift.
"""
import os
import json
import xml.etree.ElementTree as ET
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r"C:\Users\Scott\CLAUDE CODE\hytek-rfy-tools\tmp_detailer_test\multi-job"
OUT = r"C:\Users\Scott\CLAUDE CODE\hytek-rfy-tools\tmp_detailer_test\MULTI-JOB-COMPARE-v2.xlsx"


def parse_inner_xml(path):
    tree = ET.parse(path)
    root = tree.getroot()
    out = {}
    for plan in root.iter("plan"):
        for frame in plan.findall("frame"):
            fname = frame.get("name", "?")
            stick_map = {}
            for stick in frame.findall("stick"):
                sname = stick.get("name", "?")
                ops = []
                tooling = stick.find("tooling")
                if tooling is not None:
                    for op in tooling:
                        ops.append({
                            "tag": op.tag, "type": op.get("type", "?"),
                            "pos": op.get("pos"),
                            "start": op.get("startPos"), "end": op.get("endPos"),
                        })
                stick_map[sname] = ops
            out[fname] = stick_map
    return out


def op_to_str(op):
    tag = op["tag"].replace("-tool", "")
    t = op["type"]
    if op.get("pos") is not None:
        return f"{t} @ {op['pos']} ({tag})"
    if op.get("start") is not None and op.get("end") is not None:
        return f"{t} {op['start']}..{op['end']} ({tag})"
    return f"{t} ({tag})"


def op_pos(op):
    return float(op.get("pos") or op.get("start") or 0)


def categorise_drift(d):
    if d <= 0.5: return "exact"
    if d <= 5:   return "drift-small"
    if d <= 30:  return "drift-medium"
    return "drift-large"


def pair_ops_nn(ref_ops, codec_ops):
    """Nearest-neighbour pairing per (type, tag) group.
    Returns list of (ref_op, codec_op, status, drift_mm)."""
    # Group by (type, tag)
    def key(op): return (op["type"], op["tag"])
    ref_groups = {}
    codec_groups = {}
    for r in ref_ops: ref_groups.setdefault(key(r), []).append(r)
    for c in codec_ops: codec_groups.setdefault(key(c), []).append(c)

    out = []
    all_keys = set(ref_groups) | set(codec_groups)
    for k in all_keys:
        rs = sorted(ref_groups.get(k, []), key=op_pos)
        cs = sorted(codec_groups.get(k, []), key=op_pos)
        # Greedy nearest pairing: for each ref op, take closest unmatched codec op
        used = [False] * len(cs)
        for r in rs:
            rp = op_pos(r)
            best_i = -1; best_d = 1e9
            for i, c in enumerate(cs):
                if used[i]: continue
                d = abs(op_pos(c) - rp)
                if d < best_d:
                    best_d = d; best_i = i
            if best_i >= 0:
                used[best_i] = True
                out.append((r, cs[best_i], categorise_drift(best_d), best_d))
            else:
                out.append((r, None, "missing", None))
        for i, c in enumerate(cs):
            if not used[i]:
                out.append((None, c, "extra", None))
    return out


def main():
    if not os.path.isdir(BASE):
        print(f"ERROR: {BASE} not found")
        return

    jobs = []
    for jobnum in sorted(os.listdir(BASE)):
        jp = os.path.join(BASE, jobnum)
        if not os.path.isdir(jp): continue
        meta_path = os.path.join(jp, "meta.json")
        if not os.path.isfile(meta_path): continue
        with open(meta_path) as f: meta = json.load(f)
        ref_xml = next((os.path.join(jp, f) for f in os.listdir(jp) if f.endswith(".detailer-ref.xml")), None)
        codec_xml = next((os.path.join(jp, f) for f in os.listdir(jp) if f.endswith(".codec.xml")), None)
        if ref_xml and codec_xml:
            jobs.append({"jobnum": jobnum, "meta": meta, "ref_xml": ref_xml, "codec_xml": codec_xml})

    print(f"Building v2 spreadsheet for {len(jobs)} jobs:")
    wb = Workbook()
    wb.remove(wb.active)

    H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    H_FILL = PatternFill("solid", start_color="305496")
    SECT_FONT = Font(name="Arial", bold=True, size=11)
    SECT_FILL = PatternFill("solid", start_color="D9E1F2")
    BODY = Font(name="Arial", size=10)
    EXACT_FILL = PatternFill("solid", start_color="C6EFCE")     # light green
    EXACT_FONT = Font(name="Arial", color="006100", size=10)
    SMALL_FILL = PatternFill("solid", start_color="E2EFDA")     # paler green-ish
    SMALL_FONT = Font(name="Arial", color="385723", size=10)
    MED_FILL   = PatternFill("solid", start_color="FFF2CC")     # yellow
    MED_FONT   = Font(name="Arial", color="806000", size=10, italic=True)
    LARGE_FILL = PatternFill("solid", start_color="FFCC99")     # orange
    LARGE_FONT = Font(name="Arial", color="C65911", bold=True, size=10)
    MISS_FILL  = PatternFill("solid", start_color="FFC7CE")     # red
    MISS_FONT  = Font(name="Arial", color="9C0006", bold=True, size=10)
    LINK_FONT  = Font(name="Arial", color="0563C1", underline="single", size=10)
    thin = Side(border_style="thin", color="BFBFBF")
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_STYLE = {
        "exact":        ("MATCH",          EXACT_FILL, EXACT_FONT),
        "drift-small":  ("~match (≤5mm)",  SMALL_FILL, SMALL_FONT),
        "drift-medium": ("DRIFT (≤30mm)",  MED_FILL,   MED_FONT),
        "drift-large":  ("BIG DRIFT",      LARGE_FILL, LARGE_FONT),
        "missing":      ("MISSING",        MISS_FILL,  MISS_FONT),
        "extra":        ("EXTRA",          MISS_FILL,  MISS_FONT),
    }

    summary_data = []
    aggregate = Counter()  # (op_type, status) → count

    for job in jobs:
        jobnum = job["jobnum"]
        meta = job["meta"]
        plan_name = meta["planName"]

        print(f"  {jobnum} {plan_name}...")
        ref = parse_inner_xml(job["ref_xml"])
        codec = parse_inner_xml(job["codec_xml"])

        ws = wb.create_sheet(title=jobnum)
        ws.cell(row=1, column=1, value=f"{jobnum} — {meta['builder']} — {plan_name}").font = Font(name="Arial", bold=True, size=14)
        ws.merge_cells("A1:G1")
        for i, (label, val) in enumerate([
            ("Input XML (Y:):", meta["xmlPath"]),
            ("Detailer ref RFY (Y:):", meta["refRfyPath"]),
        ], start=3):
            ws.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
            c = ws.cell(row=i, column=2, value=val); c.font = LINK_FONT
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
            if val.startswith(("Y:", "C:")):
                c.hyperlink = "file:///" + val.replace("\\", "/")

        table_start = 6
        for col, h in enumerate(["Frame", "Stick", "Op #", "Detailer (ref)", "Codec (ours)", "Status", "Drift mm"], 1):
            c = ws.cell(row=table_start, column=col, value=h)
            c.font = H_FONT; c.fill = H_FILL; c.border = BOX
            c.alignment = Alignment(horizontal="center")
        row = table_start + 1

        cnt = Counter()
        type_status = Counter()  # (op_type, status) → count for this job

        for fname in ref.keys():
            ref_sticks = ref[fname]
            codec_sticks = codec.get(fname, {})
            all_sticks = list(ref_sticks.keys())
            for s in codec_sticks.keys():
                if s not in all_sticks: all_sticks.append(s)

            c = ws.cell(row=row, column=1, value=f"FRAME {fname}")
            c.font = SECT_FONT; c.fill = SECT_FILL; c.border = BOX
            for col in range(2, 8):
                ws.cell(row=row, column=col).fill = SECT_FILL
                ws.cell(row=row, column=col).border = BOX
            row += 1

            for sname in all_sticks:
                r_ops = ref_sticks.get(sname, [])
                c_ops = codec_sticks.get(sname, [])

                # Stick subhead
                c2 = ws.cell(row=row, column=2, value=f"{sname}  (D:{len(r_ops)} / C:{len(c_ops)})")
                c2.font = Font(name="Arial", bold=True, size=10); c2.border = BOX
                c2.fill = PatternFill("solid", start_color="F2F2F2")
                for col in [1, 3, 4, 5, 6, 7]:
                    ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="F2F2F2")
                    ws.cell(row=row, column=col).border = BOX
                row += 1

                paired = pair_ops_nn(r_ops, c_ops)
                # Sort for display: by ref pos when present, else codec pos
                def sort_key(p):
                    r, c, _, _ = p
                    return (r["type"] if r else c["type"], op_pos(r) if r else op_pos(c))
                paired.sort(key=sort_key)

                for i, (r_op, c_op, status, drift) in enumerate(paired, 1):
                    cnt[status] += 1
                    op_t = (r_op or c_op)["type"]
                    type_status[(op_t, status)] += 1
                    aggregate[(op_t, status)] += 1

                    label, fill, font = STATUS_STYLE[status]

                    ws.cell(row=row, column=1).border = BOX
                    ws.cell(row=row, column=2).border = BOX
                    ws.cell(row=row, column=3, value=i).border = BOX
                    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                    rc = ws.cell(row=row, column=4, value=op_to_str(r_op) if r_op else "(none)")
                    cc = ws.cell(row=row, column=5, value=op_to_str(c_op) if c_op else "(none)")
                    rc.font = BODY; cc.font = BODY; rc.border = BOX; cc.border = BOX

                    sc = ws.cell(row=row, column=6, value=label)
                    sc.font = font; sc.fill = fill; sc.border = BOX
                    sc.alignment = Alignment(horizontal="center")
                    if status in ("missing", "extra", "drift-large", "drift-medium"):
                        rc.fill = fill; cc.fill = fill
                    drift_str = f"{drift:.1f}" if drift is not None else ""
                    dc = ws.cell(row=row, column=7, value=drift_str)
                    dc.border = BOX
                    dc.alignment = Alignment(horizontal="center")
                    if status == "drift-large":
                        dc.font = LARGE_FONT

                    row += 1

                row += 1

        # Per-job summary
        row += 1
        ws.cell(row=row, column=1, value="JOB STATUS BREAKDOWN").font = Font(name="Arial", bold=True, size=12)
        row += 1
        for status, count in [("exact", cnt["exact"]), ("drift-small", cnt["drift-small"]),
                              ("drift-medium", cnt["drift-medium"]), ("drift-large", cnt["drift-large"]),
                              ("missing", cnt["missing"]), ("extra", cnt["extra"])]:
            label, _, _ = STATUS_STYLE[status]
            ws.cell(row=row, column=4, value=label).font = Font(name="Arial", bold=True)
            ws.cell(row=row, column=5, value=count)
            row += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 6
        ws.column_dimensions["D"].width = 36
        ws.column_dimensions["E"].width = 36
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 10
        ws.freeze_panes = f"A{table_start + 1}"

        summary_data.append({"jobnum": jobnum, "builder": meta["builder"], "plan": plan_name, "cnt": dict(cnt), "type_status": dict(type_status)})

    # ==== Summary sheet ====
    s = wb.create_sheet(title="_Summary", index=0)
    s.cell(row=1, column=1, value="HYTEK Codec vs Detailer — multi-job, nearest-neighbour matched").font = Font(name="Arial", bold=True, size=14)
    s.merge_cells("A1:I1")

    s.cell(row=3, column=1, value="Per-job status counts (each row = one plan)").font = Font(name="Arial", bold=True, size=12)
    headers = ["Job", "Builder", "Plan", "Exact", "Drift ≤5mm", "Drift ≤30mm", "Big Drift", "Missing", "Extra", "True match %", "True+drift %"]
    for col, h in enumerate(headers, 1):
        c = s.cell(row=4, column=col, value=h)
        c.font = H_FONT; c.fill = H_FILL; c.border = BOX
        c.alignment = Alignment(horizontal="center")

    row = 5
    grand = Counter()
    for sd in summary_data:
        cnt = Counter(sd["cnt"])
        grand.update(cnt)
        total_d = cnt["exact"] + cnt["drift-small"] + cnt["drift-medium"] + cnt["drift-large"] + cnt["missing"]
        true_match = cnt["exact"] + cnt["drift-small"]
        type_match = true_match + cnt["drift-medium"] + cnt["drift-large"]
        s.cell(row=row, column=1, value=sd["jobnum"]).border = BOX
        s.cell(row=row, column=2, value=sd["builder"]).border = BOX
        s.cell(row=row, column=3, value=sd["plan"]).border = BOX
        s.cell(row=row, column=4, value=cnt["exact"]).border = BOX
        s.cell(row=row, column=5, value=cnt["drift-small"]).border = BOX
        s.cell(row=row, column=6, value=cnt["drift-medium"]).border = BOX
        s.cell(row=row, column=7, value=cnt["drift-large"]).border = BOX
        s.cell(row=row, column=8, value=cnt["missing"]).border = BOX
        s.cell(row=row, column=9, value=cnt["extra"]).border = BOX
        s.cell(row=row, column=10, value=f"{100*true_match/total_d:.1f}%" if total_d else "-").border = BOX
        s.cell(row=row, column=11, value=f"{100*type_match/total_d:.1f}%" if total_d else "-").border = BOX
        row += 1

    # Grand totals
    total_d = grand["exact"] + grand["drift-small"] + grand["drift-medium"] + grand["drift-large"] + grand["missing"]
    true_match = grand["exact"] + grand["drift-small"]
    type_match = true_match + grand["drift-medium"] + grand["drift-large"]
    s.cell(row=row, column=1, value="TOTAL").font = SECT_FONT
    s.cell(row=row, column=4, value=grand["exact"]).font = SECT_FONT
    s.cell(row=row, column=5, value=grand["drift-small"]).font = SECT_FONT
    s.cell(row=row, column=6, value=grand["drift-medium"]).font = SECT_FONT
    s.cell(row=row, column=7, value=grand["drift-large"]).font = SECT_FONT
    s.cell(row=row, column=8, value=grand["missing"]).font = SECT_FONT
    s.cell(row=row, column=9, value=grand["extra"]).font = SECT_FONT
    s.cell(row=row, column=10, value=f"{100*true_match/total_d:.1f}%" if total_d else "-").font = SECT_FONT
    s.cell(row=row, column=11, value=f"{100*type_match/total_d:.1f}%" if total_d else "-").font = SECT_FONT
    for col in range(1, 12):
        s.cell(row=row, column=col).fill = SECT_FILL
        s.cell(row=row, column=col).border = BOX
    row += 3

    # Aggregate by op-type
    s.cell(row=row, column=1, value="Aggregate by op-type — what's the gap shape per rule?").font = Font(name="Arial", bold=True, size=12)
    row += 1
    cols = ["Op type", "Exact", "Drift ≤5mm", "Drift ≤30mm", "Big Drift", "Missing", "Extra"]
    for col, h in enumerate(cols, 1):
        c = s.cell(row=row, column=col, value=h)
        c.font = H_FONT; c.fill = H_FILL; c.border = BOX
    row += 1
    op_types = sorted(set(t for (t, _), _ in aggregate.items()),
                     key=lambda t: -sum(aggregate.get((t, st), 0) for st in ("exact","drift-small","drift-medium","drift-large","missing","extra")))
    for op_t in op_types:
        s.cell(row=row, column=1, value=op_t).border = BOX
        for i, st in enumerate(["exact", "drift-small", "drift-medium", "drift-large", "missing", "extra"], 2):
            v = aggregate.get((op_t, st), 0)
            c = s.cell(row=row, column=i, value=v); c.border = BOX
            if v > 0 and st in ("missing", "extra"):
                c.fill = MISS_FILL; c.font = MISS_FONT
            elif v > 0 and st == "drift-large":
                c.fill = LARGE_FILL; c.font = LARGE_FONT
            elif v > 0 and st == "drift-medium":
                c.fill = MED_FILL; c.font = MED_FONT
        row += 1

    # Legend
    row += 2
    s.cell(row=row, column=1, value="Status legend:").font = Font(name="Arial", bold=True, size=11)
    row += 1
    for st in ["exact", "drift-small", "drift-medium", "drift-large", "missing", "extra"]:
        label, fill, font = STATUS_STYLE[st]
        c = s.cell(row=row, column=1, value=label)
        c.fill = fill; c.font = font; c.border = BOX
        desc = {
            "exact": "Bit-perfect — same op type at exact same position",
            "drift-small": "Same type, position within 5mm (effectively a match — sub-mm rounding)",
            "drift-medium": "Same type, position 5-30mm off (NEW: was double-counted as missing+extra in v1)",
            "drift-large": "Same type, but position is wildly different (>30mm)",
            "missing": "Detailer emits this op TYPE; codec emits NONE of this type at this position",
            "extra": "Codec emits this op TYPE; Detailer emits NONE",
        }[st]
        s.cell(row=row, column=2, value=desc).font = Font(name="Arial", italic=True)
        s.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        row += 1

    s.column_dimensions["A"].width = 16
    s.column_dimensions["B"].width = 28
    s.column_dimensions["C"].width = 24
    for c in "DEFGHIJK":
        s.column_dimensions[c].width = 13
    s.freeze_panes = "A5"

    wb.save(OUT)

    # Print headlines
    print()
    print(f"Saved: {OUT}")
    print()
    print(f"GRAND TOTAL across {len(summary_data)} jobs:")
    print(f"  Exact match:       {grand['exact']:>6,}  ({100*grand['exact']/total_d:.1f}%)")
    print(f"  Drift ≤5mm:        {grand['drift-small']:>6,}  ({100*grand['drift-small']/total_d:.1f}%)")
    print(f"  Drift ≤30mm:       {grand['drift-medium']:>6,}  ({100*grand['drift-medium']/total_d:.1f}%)")
    print(f"  Drift >30mm:       {grand['drift-large']:>6,}  ({100*grand['drift-large']/total_d:.1f}%)")
    print(f"  TRULY missing:     {grand['missing']:>6,}  ({100*grand['missing']/total_d:.1f}%)")
    print(f"  TRULY extra:       {grand['extra']:>6,}")
    print()
    print(f"  TRUE match (exact+drift≤5mm):              {100*true_match/total_d:.1f}%")
    print(f"  TYPE match (any drift counted as found):   {100*type_match/total_d:.1f}%")


if __name__ == "__main__":
    main()
