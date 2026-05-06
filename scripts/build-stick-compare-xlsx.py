"""Build a side-by-side stick-by-stick operations comparison spreadsheet.

For each stick in HG260017 GF-LBW-70.075, lists Detailer's operations vs our
codec's operations. Differences highlighted red.
"""
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REF_XML = r"C:\Users\Scott\CLAUDE CODE\hytek-rfy-tools\tmp_detailer_test\COMPARE-2-detailer-now.xml"
CODEC_XML = r"C:\Users\Scott\CLAUDE CODE\hytek-rfy-tools\tmp_detailer_test\COMPARE-3-codec.xml"
OUT_PATH = r"C:\Users\Scott\CLAUDE CODE\hytek-rfy-tools\tmp_detailer_test\STICK-COMPARE-HG260017-GF-LBW-70.075.xlsx"


def parse_inner_xml(path):
    """Parse Detailer's inner schedule XML.
    Returns dict: {frame_name: {stick_name: [{type, kind, pos|start|end, ...}, ...]}}
    """
    tree = ET.parse(path)
    root = tree.getroot()  # <schedule>
    out = {}
    for plan in root.iter("plan"):
        for frame in plan.findall("frame"):
            fname = frame.get("name", "?")
            stick_map = {}
            for stick in frame.findall("stick"):
                sname = stick.get("name", "?")
                ops = []
                tooling = stick.find("tooling")
                if tooling is not None:
                    for op in tooling:
                        tag = op.tag  # 'point-tool', 'spanned-tool', 'start-tool', 'end-tool'
                        ops.append({
                            "tag": tag,
                            "type": op.get("type", "?"),
                            "pos": op.get("pos"),
                            "start": op.get("startPos"),
                            "end": op.get("endPos"),
                        })
                # Sort by tag-type-pos for stable comparison
                stick_map[sname] = ops
            out[fname] = stick_map
    return out


def op_to_str(op):
    """Format an op as a short string for display."""
    tag = op["tag"].replace("-tool", "")  # point/spanned/start/end
    t = op["type"]
    if op.get("pos") is not None:
        return f"{t} @ {op['pos']} ({tag})"
    if op.get("start") is not None and op.get("end") is not None:
        return f"{t} {op['start']}..{op['end']} ({tag})"
    return f"{t} ({tag})"


def op_key(op):
    """Stable sort key."""
    p = op.get("pos") or op.get("start") or "0"
    try:
        p = float(p)
    except Exception:
        p = 0
    return (op["type"], op["tag"], p)


def main():
    print(f"Parsing Detailer XML: {REF_XML}")
    ref = parse_inner_xml(REF_XML)
    print(f"Parsing codec XML:    {CODEC_XML}")
    codec = parse_inner_xml(CODEC_XML)

    wb = Workbook()
    ws = wb.active
    ws.title = "HG260017 GF-LBW-70.075"

    # Styles
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="305496")
    section_font = Font(name="Arial", bold=True, size=11)
    section_fill = PatternFill("solid", start_color="D9E1F2")
    diff_fill = PatternFill("solid", start_color="FFCCCC")  # light red
    diff_font = Font(name="Arial", color="C00000", bold=True)
    detailer_only_fill = PatternFill("solid", start_color="FFE699")  # yellow — missing from codec
    codec_only_fill = PatternFill("solid", start_color="FFC7CE")  # red — extra in codec
    body_font = Font(name="Arial", size=10)
    thin = Side(border_style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    headers = ["Frame", "Stick", "Op #", "Detailer (ref)", "Codec (ours)", "Match?"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = box

    row = 2
    total_match = 0
    total_detailer = 0
    total_codec = 0

    # Iterate frames in deterministic order (Detailer's order)
    for fname in ref.keys():
        ref_sticks = ref[fname]
        codec_sticks = codec.get(fname, {})
        # All sticks in either side
        all_stick_names = list(ref_sticks.keys())
        for s in codec_sticks.keys():
            if s not in all_stick_names:
                all_stick_names.append(s)
        for sname in all_stick_names:
            ref_ops = sorted(ref_sticks.get(sname, []), key=op_key)
            codec_ops = sorted(codec_sticks.get(sname, []), key=op_key)
            total_detailer += len(ref_ops)
            total_codec += len(codec_ops)

            # Stick header row
            stick_summary = f"{sname}  (Detailer: {len(ref_ops)} ops, Codec: {len(codec_ops)} ops)"
            c = ws.cell(row=row, column=1, value=fname)
            c.font = section_font; c.fill = section_fill; c.border = box
            c = ws.cell(row=row, column=2, value=stick_summary)
            c.font = section_font; c.fill = section_fill; c.border = box
            for col in range(3, 7):
                ws.cell(row=row, column=col).fill = section_fill
                ws.cell(row=row, column=col).border = box
            row += 1

            # Pair-up ops: greedy matching by (type, tag, approx-pos)
            # For each Detailer op, find matching codec op (consume it)
            codec_pool = list(codec_ops)
            paired_rows = []
            for r_op in ref_ops:
                best_idx = -1
                for i, c_op in enumerate(codec_pool):
                    if c_op["type"] == r_op["type"] and c_op["tag"] == r_op["tag"]:
                        # match positions within 5mm tolerance
                        rp = float(r_op.get("pos") or r_op.get("start") or 0)
                        cp = float(c_op.get("pos") or c_op.get("start") or 0)
                        if abs(rp - cp) < 5:
                            best_idx = i
                            break
                if best_idx >= 0:
                    matched = codec_pool.pop(best_idx)
                    paired_rows.append((r_op, matched, "match"))
                    total_match += 1
                else:
                    # Detailer-only (missing from codec)
                    paired_rows.append((r_op, None, "missing"))
            # Whatever's left in codec_pool is extra in codec
            for c_op in codec_pool:
                paired_rows.append((None, c_op, "extra"))

            # Emit rows
            for op_idx, (r_op, c_op, status) in enumerate(paired_rows, 1):
                ws.cell(row=row, column=1, value="").border = box
                ws.cell(row=row, column=2, value="").border = box
                ws.cell(row=row, column=3, value=op_idx).border = box
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

                ref_str = op_to_str(r_op) if r_op else ""
                codec_str = op_to_str(c_op) if c_op else ""
                rc = ws.cell(row=row, column=4, value=ref_str); rc.font = body_font; rc.border = box
                cc = ws.cell(row=row, column=5, value=codec_str); cc.font = body_font; cc.border = box

                if status == "match":
                    # Slight value difference detection: if r_op pos != c_op pos
                    rp = (r_op.get("pos") or f"{r_op.get('start')}..{r_op.get('end')}")
                    cp = (c_op.get("pos") or f"{c_op.get('start')}..{c_op.get('end')}")
                    if rp != cp:
                        # Match by type but different position
                        ws.cell(row=row, column=6, value="~match (pos diff)").font = Font(name="Arial", italic=True, color="806000")
                        rc.fill = PatternFill("solid", start_color="FFF2CC")
                        cc.fill = PatternFill("solid", start_color="FFF2CC")
                    else:
                        ws.cell(row=row, column=6, value="match").font = Font(name="Arial", color="385723")
                elif status == "missing":
                    cc.fill = detailer_only_fill
                    cc.value = "(none — codec missing this op)"
                    cc.font = diff_font
                    rc.fill = detailer_only_fill
                    ws.cell(row=row, column=6, value="MISSING").font = diff_font
                    ws.cell(row=row, column=6).fill = diff_fill
                else:  # extra
                    rc.fill = codec_only_fill
                    rc.value = "(none — codec emits extra op)"
                    rc.font = diff_font
                    cc.fill = codec_only_fill
                    ws.cell(row=row, column=6, value="EXTRA").font = diff_font
                    ws.cell(row=row, column=6).fill = diff_fill

                ws.cell(row=row, column=6).border = box
                ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
                row += 1

            # Blank separator row
            row += 1

    # Summary at the bottom
    summary_row = row + 1
    ws.cell(row=summary_row, column=1, value="TOTALS").font = Font(name="Arial", bold=True, size=12)
    summary_row += 1
    ws.cell(row=summary_row, column=4, value="Detailer ops:").font = Font(name="Arial", bold=True)
    ws.cell(row=summary_row, column=5, value=total_detailer)
    summary_row += 1
    ws.cell(row=summary_row, column=4, value="Codec ops:").font = Font(name="Arial", bold=True)
    ws.cell(row=summary_row, column=5, value=total_codec)
    summary_row += 1
    ws.cell(row=summary_row, column=4, value="Matched:").font = Font(name="Arial", bold=True)
    ws.cell(row=summary_row, column=5, value=total_match)
    summary_row += 1
    ws.cell(row=summary_row, column=4, value="Match %:").font = Font(name="Arial", bold=True)
    if total_detailer > 0:
        ws.cell(row=summary_row, column=5, value=f"{100*total_match/total_detailer:.1f}%")
    summary_row += 1
    ws.cell(row=summary_row, column=4, value="Codec missing:").font = Font(name="Arial", bold=True)
    ws.cell(row=summary_row, column=5, value=total_detailer - total_match)
    summary_row += 1
    ws.cell(row=summary_row, column=4, value="Codec extra:").font = Font(name="Arial", bold=True)
    ws.cell(row=summary_row, column=5, value=total_codec - total_match)

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 38
    ws.column_dimensions["F"].width = 18

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")
    print(f"  Detailer total ops: {total_detailer}")
    print(f"  Codec total ops:    {total_codec}")
    print(f"  Matched:            {total_match}")
    if total_detailer:
        print(f"  Match %:            {100*total_match/total_detailer:.1f}%")


if __name__ == "__main__":
    main()
